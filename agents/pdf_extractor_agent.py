"""
PDF Data Extraction Agent
Reads scanned/image-based PDFs using Groq vision model,
extracts all structured data, and writes to a formatted Excel file.
"""
import sys, os
import sys
sys.set_int_max_str_digits(100000)  # fix large integer conversion limit
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import argparse
import base64
import io
import json
import time
import re
from pathlib import Path

import fitz                      # pymupdf
from PIL import Image
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from groq import Groq
from dotenv import load_dotenv

load_dotenv()

client = Groq(api_key=os.getenv("GROQ_API_KEY"))
VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"  # Groq vision model


# ── PDF → base64 images ────────────────────────────────────────────────────

def pdf_page_to_base64(pdf_path: str, page_num: int, dpi: int = 72) -> str:
    """Convert a PDF page to a base64-encoded JPEG for vision API (compressed)."""
    doc = fitz.open(pdf_path)
    page = doc[page_num]
    scale = dpi / 72
    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat)
    img = Image.open(io.BytesIO(pix.tobytes("png")))
    # Resize to max 1568px on longest side (Groq vision limit)
    max_size = 1568
    ratio = min(max_size / img.width, max_size / img.height, 1.0)
    if ratio < 1.0:
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.LANCZOS)
    # Save as JPEG with quality 85 to reduce size
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=85, optimize=True)
    doc.close()
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def get_page_count(pdf_path: str) -> int:
    doc = fitz.open(pdf_path)
    n = len(doc)
    doc.close()
    return n


# ── Vision extraction ──────────────────────────────────────────────────────

EXTRACTION_PROMPT = """You are a precise data extraction specialist. 
Analyze this page from the NASBA (National Association of State Boards of Accountancy) 
2020 Performance Book.

Extract ALL data you can see with 100% accuracy. Return a JSON object with this structure:

{
  "page_title": "title of this page/section",
  "page_type": "table|chart|text|mixed",
  "tables": [
    {
      "table_title": "name of the table",
      "headers": ["col1", "col2", ...],
      "rows": [
        ["val1", "val2", ...],
        ...
      ]
    }
  ],
  "key_metrics": [
    {"metric": "name", "value": "value", "context": "any notes"}
  ],
  "text_content": "any important narrative text not in tables",
  "notes": "any footnotes or special annotations"
}

Rules:
- Extract EVERY number, percentage, and name exactly as shown
- Preserve all decimal places and formatting
- If a cell is empty, use null
- If you see N/A or dashes, preserve them as strings
- For charts/graphs, extract the underlying data values if visible
- Return ONLY valid JSON, no markdown, no explanation
"""


def extract_page_data(pdf_path: str, page_num: int, retries: int = 3) -> dict:
    """Extract structured data from a single PDF page using vision model."""
    for attempt in range(retries):
        try:
            b64 = pdf_page_to_base64(pdf_path, page_num)
            response = client.chat.completions.create(
                model=VISION_MODEL,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text",    "text": EXTRACTION_PROMPT},
                        {"type": "image_url", "image_url": {
                            "url": f"data:image/jpeg;base64,{b64}"
                        }}
                    ]
                }],
                max_tokens=4096,
                temperature=0.1,
            )
            raw = response.choices[0].message.content.strip()
            # Strip markdown code fences if present
            raw = re.sub(r"^```json\s*", "", raw)
            raw = re.sub(r"^```\s*",     "", raw)
            raw = re.sub(r"\s*```$",     "", raw)
            return json.loads(raw)
        except json.JSONDecodeError:
            # Try to extract JSON from response
            match = re.search(r'\{.*\}', raw, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group())
                except:
                    pass
            print(f"  ⚠ Page {page_num+1}: JSON parse failed (attempt {attempt+1})")
            time.sleep(2)
        except Exception as e:
            if "429" in str(e):
                wait = (attempt + 1) * 20
                print(f"  ⏳ Rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  ✗ Page {page_num+1} error: {e}")
                return {}
    print(f"  ✗ Page {page_num+1}: giving up after {retries} attempts")
    return {}


# ── Excel writer ───────────────────────────────────────────────────────────

# Color palette
DARK_BG     = "1E293B"
HEADER_BG   = "6366F1"
ALT_ROW     = "111827"
ACCENT      = "06B6D4"
TEXT_LIGHT  = "F1F5F9"
TEXT_MUTED  = "94A3B8"
BORDER_COL  = "334155"
WHITE       = "FFFFFF"
GREEN       = "10B981"
YELLOW      = "F59E0B"


def make_border(color=BORDER_COL):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_cell(cell, text, bg=HEADER_BG, font_color=WHITE, size=11, bold=True):
    cell.value = text
    cell.font  = Font(name="Calibri", bold=bold, color=font_color, size=size)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def style_data_cell(cell, value, row_idx=0, align="left"):
    cell.value = value
    bg = ALT_ROW if row_idx % 2 == 0 else "0F172A"
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.font  = Font(name="Calibri", color=TEXT_LIGHT, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()


def style_title_cell(cell, text, bg=DARK_BG):
    cell.value = text
    cell.font  = Font(name="Calibri", bold=True, color=ACCENT, size=13)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_page_to_sheet(wb: openpyxl.Workbook, page_num: int, data: dict):
    """Write extracted page data to a dedicated Excel sheet."""
    title = data.get("page_title", f"Page {page_num+1}")
    # Sanitize sheet name
    sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(title))[:31] or f"Page_{page_num+1}"

    # Avoid duplicate sheet names
    existing = [s.title for s in wb.worksheets]
    if sheet_name in existing:
        sheet_name = f"{sheet_name[:27]}_{page_num+1}"

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = HEADER_BG

    row = 1

    # ── Page title banner ──
    ws.row_dimensions[row].height = 36
    ws.merge_cells(f"A{row}:J{row}")
    style_title_cell(ws[f"A{row}"], f"  {title}")
    ws[f"A{row}"].fill = PatternFill("solid", fgColor="0F172A")
    row += 1

    # ── Page type badge ──
    ws.row_dimensions[row].height = 20
    ws[f"A{row}"].value = f"Page {page_num+1}  ·  Type: {data.get('page_type','—').upper()}"
    ws[f"A{row}"].font  = Font(name="Calibri", color=TEXT_MUTED, size=9, italic=True)
    ws[f"A{row}"].fill  = PatternFill("solid", fgColor="0A0A0F")
    row += 2

    # ── Tables ────────────────────────────────────────────────────────────
    for tbl in data.get("tables", []):
        headers = tbl.get("headers", [])
        rows    = tbl.get("rows", [])
        if not headers and not rows:
            continue

        # Table title
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"A{row}:{get_column_letter(max(len(headers),1)+1)}{row}")
        ws[f"A{row}"].value = f"  📊  {tbl.get('table_title','Table')}"
        ws[f"A{row}"].font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
        ws[f"A{row}"].fill  = PatternFill("solid", fgColor="1E293B")
        ws[f"A{row}"].alignment = Alignment(vertical="center")
        row += 1

        # Headers
        ws.row_dimensions[row].height = 30
        for ci, h in enumerate(headers, 1):
            style_header_cell(ws.cell(row, ci), h)
        row += 1

        # Data rows
        for ri, data_row in enumerate(rows):
            ws.row_dimensions[row].height = 22
            for ci, val in enumerate(data_row, 1):
                align = "right" if isinstance(val, (int, float)) else "left"
                # Try to detect numeric values
                display = val
                if val is not None:
                    try:
                        display = float(str(val).replace(",","").replace("%",""))
                        align = "right"
                    except:
                        display = val
                style_data_cell(ws.cell(row, ci), display, ri, align)
            row += 1

        row += 2  # gap between tables

    # ── Key metrics ───────────────────────────────────────────────────────
    metrics = data.get("key_metrics", [])
    if metrics:
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].value = "  📈  Key Metrics"
        ws[f"A{row}"].font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
        ws[f"A{row}"].fill  = PatternFill("solid", fgColor="1E293B")
        ws[f"A{row}"].alignment = Alignment(vertical="center")
        row += 1

        for hi, h in enumerate(["Metric", "Value", "Context"], 1):
            style_header_cell(ws.cell(row, hi), h)
        row += 1

        for ri, m in enumerate(metrics):
            style_data_cell(ws.cell(row, 1), m.get("metric",""), ri)
            style_data_cell(ws.cell(row, 2), m.get("value",""),  ri, "right")
            style_data_cell(ws.cell(row, 3), m.get("context",""),ri)
            row += 1

        row += 2

    # ── Text content ──────────────────────────────────────────────────────
    text = data.get("text_content", "")
    if text and text.strip():
        ws.row_dimensions[row].height = 22
        ws[f"A{row}"].value = "  📝  Notes & Narrative"
        ws[f"A{row}"].font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
        ws[f"A{row}"].fill  = PatternFill("solid", fgColor="1E293B")
        row += 1
        ws.merge_cells(f"A{row}:J{row+4}")
        ws[f"A{row}"].value = text
        ws[f"A{row}"].font  = Font(name="Calibri", color=TEXT_MUTED, size=10)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"A{row}"].fill  = PatternFill("solid", fgColor=ALT_ROW)
        row += 6

    # ── Auto-fit columns ──────────────────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    # Freeze top rows
    ws.freeze_panes = "A3"


def create_summary_sheet(wb: openpyxl.Workbook, all_data: list):
    """Create a summary/index sheet as the first sheet."""
    ws = wb.create_sheet("📋 Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "NASBA 2020 Performance Book — Extracted Data"
    ws["A1"].font  = Font(name="Calibri", bold=True, color=WHITE, size=16)
    ws["A1"].fill  = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Extracted {len(all_data)} pages  ·  Generated by AI Agent"
    ws["A2"].font  = Font(name="Calibri", color=TEXT_MUTED, size=10, italic=True)
    ws["A2"].fill  = PatternFill("solid", fgColor="0A0A0F")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    row = 4
    for hi, h in enumerate(["Page", "Title", "Type", "Tables", "Metrics", "Has Text"], 1):
        style_header_cell(ws.cell(row, hi), h)
    ws.row_dimensions[row].height = 28
    row += 1

    for i, data in enumerate(all_data):
        ws.row_dimensions[row].height = 22
        style_data_cell(ws.cell(row, 1), i+1, i, "center")
        style_data_cell(ws.cell(row, 2), data.get("page_title","—"), i)
        style_data_cell(ws.cell(row, 3), data.get("page_type","—").upper(), i, "center")
        style_data_cell(ws.cell(row, 4), len(data.get("tables",[])), i, "center")
        style_data_cell(ws.cell(row, 5), len(data.get("key_metrics",[])), i, "center")
        style_data_cell(ws.cell(row, 6), "Yes" if str(data.get("text_content") or "").strip() else "No", i, "center")
        row += 1

    for col, width in zip("ABCDEF", [8, 45, 12, 10, 10, 10]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


# ── Main extraction pipeline ───────────────────────────────────────────────

def extract_pdf_to_excel(pdf_path: str, output_path: str = None, 
                          start_page: int = 0, end_page: int = None):
    """
    Full pipeline: PDF → vision extraction → structured Excel.
    
    Args:
        pdf_path:    Path to the PDF file
        output_path: Output Excel path (auto-generated if None)
        start_page:  First page to process (0-indexed)
        end_page:    Last page to process (None = all pages)
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    total_pages = get_page_count(str(pdf_path))
    end_page    = end_page or total_pages
    pages_to_process = list(range(start_page, min(end_page, total_pages)))

    if output_path is None:
        out_dir = Path("output")
        out_dir.mkdir(exist_ok=True)
        output_path = out_dir / f"{pdf_path.stem}_extracted.xlsx"

    print(f"\n{'='*60}")
    print(f"  PDF Extraction Agent")
    print(f"  File   : {pdf_path.name}")
    print(f"  Pages  : {len(pages_to_process)} (of {total_pages})")
    print(f"  Output : {output_path}")
    print(f"{'='*60}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    all_data = []

    for idx, page_num in enumerate(pages_to_process):
        print(f"  [{idx+1}/{len(pages_to_process)}] Extracting page {page_num+1}...", end=" ", flush=True)
        data = extract_page_data(str(pdf_path), page_num)

        if data:
            all_data.append(data)
            write_page_to_sheet(wb, page_num, data)
            title = data.get("page_title", "—")[:40]
            tables = len(data.get("tables", []))
            metrics = len(data.get("key_metrics", []))
            print(f"✓  [{title}]  {tables} tables, {metrics} metrics")
        else:
            all_data.append({"page_title": f"Page {page_num+1}", "page_type": "empty"})
            print("⚠ No data extracted")

        # Rate limit: ~4 requests/min for vision model to be safe
        if idx < len(pages_to_process) - 1:
            time.sleep(15)

    # Create summary as first sheet
    create_summary_sheet(wb, all_data)

    wb.save(str(output_path))
    print(f"\n{'='*60}")
    print(f"  ✅ Done! Saved to: {output_path}")
    print(f"  Sheets: {len(wb.worksheets)} (1 summary + {len(wb.worksheets)-1} pages)")
    print(f"{'='*60}\n")

    return str(output_path)


# ── CLI ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PDF → Excel extraction agent")
    parser.add_argument("--pdf",   default="data/business/2020_NASBA_Performance_Book.pdf")
    parser.add_argument("--out",   default=None, help="Output Excel path")
    parser.add_argument("--start", type=int, default=0,  help="Start page (0-indexed)")
    parser.add_argument("--end",   type=int, default=None, help="End page (exclusive)")
    args = parser.parse_args()

    extract_pdf_to_excel(args.pdf, args.out, args.start, args.end)
